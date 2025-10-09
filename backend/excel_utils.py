"""Excel file processing utilities for bulk import."""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from .logging_config import get_logger

logger = get_logger(__name__)


def _find_column(sheet_data: list[dict[str, Any]], possible_names: list[str]) -> str | None:
    """Helper to find a column by multiple possible names."""
    for name in possible_names:
        if any(name in row for row in sheet_data):
            return name
    return None


def read_excel_sheets(file_bytes: bytes) -> dict[str, list[dict[str, Any]]]:
    """
    Read all sheets from an Excel file and return structured data.
    
    Args:
        file_bytes: Excel file content as bytes
        
    Returns:
        Dictionary mapping sheet names to lists of row dictionaries
    """
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        result: dict[str, list[dict[str, Any]]] = {}
        
        for sheet_name in workbook.sheetnames:
            rows = list(workbook[sheet_name].iter_rows(values_only=True))
            
            if not rows:
                logger.warning(f"Sheet '{sheet_name}' is empty")
                continue
            
            # First row is headers
            headers = [str(h).strip() if h else f"Column_{i}" for i, h in enumerate(rows[0])]
            
            # Process data rows (skip empty rows)
            sheet_data = [
                {h: str(cell).strip() if cell else "" for h, cell in zip(headers, row)}
                for row in rows[1:]
                if any(cell for cell in row)
            ]
            
            result[sheet_name] = sheet_data
            logger.info(f"Processed sheet '{sheet_name}': {len(sheet_data)} rows")
        
        workbook.close()
        return result
        
    except Exception as e:
        logger.error(f"Failed to read Excel file: {e}", exc_info=True)
        raise ValueError(f"خطا در خواندن فایل اکسل: {str(e)}")


def validate_book_sheet_data(sheet_data: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Validate and normalize book data from a sheet.
    
    Required columns: 'نام کتاب' or 'نام' or 'عنوان'
    
    Args:
        sheet_data: List of row dictionaries from a sheet
        
    Returns:
        List of validated book dictionaries with 'name' field
    """
    name_column = _find_column(sheet_data, ['نام کتاب', 'نام', 'عنوان', 'کتاب'])
    
    if not name_column:
        raise ValueError("ستون 'نام کتاب' یافت نشد. لطفاً یکی از ستون‌های 'نام کتاب'، 'نام' یا 'عنوان' را اضافه کنید.")
    
    return [
        {"name": book_name}
        for row in sheet_data
        if (book_name := row.get(name_column, "").strip())
    ]


def validate_student_sheet_data(sheet_data: list[dict[str, Any]], major: str) -> list[dict[str, Any]]:
    """
    Validate and normalize student data from a sheet.
    
    Required columns: 'نام', 'نام خانوادگی', 'پایه'
    Optional columns: 'کد ملی', 'شناسه', 'تلفن', 'شماره تماس'
    
    Args:
        sheet_data: List of row dictionaries from a sheet
        major: Major/field extracted from sheet name
        
    Returns:
        List of validated student dictionaries
    """
    # Find required columns
    first_name_col = _find_column(sheet_data, ['نام', 'نام دانش‌آموز', 'نام هنرجو'])
    last_name_col = _find_column(sheet_data, ['نام خانوادگی', 'نام‌خانوادگی', 'فامیل'])
    grade_col = _find_column(sheet_data, ['پایه', 'کلاس', 'مقطع'])
    national_id_col = _find_column(sheet_data, ['کد ملی', 'شناسه', 'کدملی', 'شناسه ملی'])
    phone_col = _find_column(sheet_data, ['تلفن', 'شماره تماس', 'شماره', 'موبایل'])
    
    if not first_name_col:
        raise ValueError("ستون 'نام' در شیت یافت نشد")
    if not last_name_col:
        raise ValueError("ستون 'نام خانوادگی' در شیت یافت نشد")
    if not grade_col:
        raise ValueError("ستون 'پایه' در شیت یافت نشد")
    
    validated_students = []
    for row in sheet_data:
        first_name = row.get(first_name_col, "").strip()
        last_name = row.get(last_name_col, "").strip()
        grade = row.get(grade_col, "").strip()
        
        if not (first_name and last_name and grade):
            continue
        
        student_data = {
            "first_name": first_name,
            "last_name": last_name,
            "grade": grade,
            "major": major,
        }
        
        # Add optional fields if present
        if national_id_col:
            national_id = row.get(national_id_col, "").strip()
            if national_id:
                student_data["national_id"] = national_id
        
        if phone_col:
            phone = row.get(phone_col, "").strip()
            if phone:
                student_data["phone_number"] = phone
        
        validated_students.append(student_data)
    
    return validated_students
